import openpyxl

# Crear un nuevo libro de trabajo de Excel
wb = openpyxl.Workbook()

# Seleccionar la hoja activa
sheet = wb.active

# Definir las cabeceras
cabeceras = ["Identificador de la pieza", "Nombre de la pieza", "Número de bulto", "Precio"]

# Agregar las cabeceras a la primera fila
for col, valor in enumerate(cabeceras, start=1):
    sheet.cell(row=1, column=col, value=valor)

# Definir los valores de las piezas
valores_piezas = []


# Agregar los valores de las piezas a las filas siguientes
for row, pieza in enumerate(valores_piezas, start=2):
    for col, valor in enumerate(pieza, start=1):
        sheet.cell(row=row, column=col, value=valor)

# Guardar el archivo de Excel
wb.save("informacion_piezas.xlsx")

print("Archivo de Excel creado exitosamente.")